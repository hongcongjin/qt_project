from openpyxl import load_workbook
from urllib import parse
from Generic_Download.General_AsyncGrab import AsyncGrab
import sqlite3
import time
import os


def parse_exl(excel_path, col_list):
    """
    :param excel_path:
    :param col_list: 需要提取的列的位置列表
    :return:
    """
    wb = load_workbook(excel_path)
    sheets = wb.sheetnames
    need_sheet = sheets[0]

    ws = wb[need_sheet]
    rows = ws.values
    ### 只提取url，并将其解码
    res_value = [[parse.unquote(row[i]) for i in col_list] for row in rows if "http" in str(row)]
    print("Excel parse Success")
    return res_value


def parse_db(dbname, table_key):
    conn = sqlite3.connect(dbname)
    s_cmd = """
    select contentUrl from "{}"
    """.format(table_key)
    cur = conn.cursor()
    cur.execute(s_cmd)
    val = cur.fetchall()
    print("db parse success")
    conn.close()
    return val


def change_from_db(org_val, savepath):
    dst_values = [[value[0], savepath + "\\" + str(num) + ".jpg"] for num, value in enumerate(org_val) if not os.path.exists(savepath + "\\" + str(num) + ".jpg")]
    return dst_values


def change_from_excl(org_values, savepath):
    for value in org_values:
        value[1] = savepath + "\\" + value[0].split('/')[-1]
    return org_values


if __name__ == '__main__':
    # exl_values = parse_exl("C:\\Users\\xcj9307\\Desktop\\0220至0224抓拍数据wcb-core.faceDetectImageDatail.xlsx", [0, 2])
    for keyword in ["Centenarian"]:   ## , "聚餐"
        #  "聚会 party", "团建", "teambuilding", "集体活动",
        # db_values = parse_db("D:\\Cralwer_Project\\Bing_for_Similar\\Bing_for_Similar.db", keyword.replace(" ", '_'))
        #
        # new_values = change_from_db(db_values, "D:\\Body\\20200907\\bing_" + keyword)
        # db_values = parse_db(r"D:\Cralwer_Project\Google_for_Similar\Generic_Download.db", keyword.replace(" ", '_'))
        db_values = parse_db(r"D:\Cralwer_Project\Google_for_Similar\Google_for_Similar.db", keyword.replace(" ", '_'))

        new_values = change_from_db(db_values, "D:\\20201125\\file\\Google__" + keyword)

        print("start time is " + time.ctime())
        AsyncGrab(new_values, 5, "content").event_loop()
        print("finish time is " + time.ctime())